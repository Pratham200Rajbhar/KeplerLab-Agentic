import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList

# Set random seed for reproducibility
np.random.seed(42)

# Generate synthetic monthly KPI data for 12 months × 6 KPIs
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
kpi_names = ['Revenue', 'Profit', 'Customers', 'ConversionRate', 'CAC', 'ROI']

# Generate synthetic data
data = {
    'Month': months,
    'Revenue': np.random.uniform(100000, 500000, 12),
    'Profit': np.random.uniform(10000, 100000, 12),
    'Customers': np.random.randint(500, 3000, 12),
    'ConversionRate': np.random.uniform(0.02, 0.12, 12),
    'CAC': np.random.uniform(20, 80, 12),
    'ROI': np.random.uniform(0.5, 3.5, 12)
}

# Create DataFrame
df = pd.DataFrame(data)

# Print dataset summary
print(f"Dataset: {df.shape[0]} rows, {df.shape[1]} columns")
print("Columns:", list(df.columns))
print("\nDescriptive Statistics:")
print(df.describe().round(2))

# Save to CSV for reference
df.to_csv('kpi_data.csv', index=False)

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "KPI Report"

# Write header row
headers = ['Month', 'Revenue', 'Profit', 'Customers', 'ConversionRate', 'CAC', 'ROI']
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Write data rows with alternating colors
light_gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
currency_format = '$#,##0.00'
percentage_format = '0.00%'
general_format = '0.00'

for row_idx, row in df.iterrows():
    row_num = row_idx + 2  # +2 because of header row and 0-indexing
    row_fill = light_gray_fill if row_idx % 2 == 1 else PatternFill()  # Alternating colors
    
    # Month
    cell = ws.cell(row=row_num, column=1, value=row['Month'])
    cell.fill = row_fill
    cell.alignment = Alignment(horizontal="center")
    
    # Revenue
    cell = ws.cell(row=row_num, column=2, value=row['Revenue'])
    cell.fill = row_fill
    cell.number_format = currency_format
    
    # Profit
    cell = ws.cell(row=row_num, column=3, value=row['Profit'])
    cell.fill = row_fill
    cell.number_format = currency_format
    
    # Customers
    cell = ws.cell(row=row_num, column=4, value=row['Customers'])
    cell.fill = row_fill
    cell.number_format = general_format
    
    # ConversionRate
    cell = ws.cell(row=row_num, column=5, value=row['ConversionRate'])
    cell.fill = row_fill
    cell.number_format = percentage_format
    
    # CAC
    cell = ws.cell(row=row_num, column=6, value=row['CAC'])
    cell.fill = row_fill
    cell.number_format = general_format
    
    # ROI
    cell = ws.cell(row=row_num, column=7, value=row['ROI'])
    cell.fill = row_fill
    cell.number_format = general_format

# Auto-adjust column widths
column_widths = []
for i, header in enumerate(headers, 1):
    # Calculate max width for each column
    max_width = len(header) + 2
    for row in range(2, len(df) + 2):
        value = ws.cell(row=row, column=i).value
        if value is not None:
            max_width = max(max_width, len(str(value)) + 2)
    column_widths.append(max_width)

for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = min(width, 25)  # Cap at 25

# Add data bars to Revenue column (column B)
data_bar_rule = DataBarRule(
    start_type='min', start_value='0',
    end_type='max', end_value=1.0,
    color="638EC7",
    showValue=True,
    minLength=0,
    maxLength=100
)
ws.conditional_formatting.add(f'B2:B{len(df)+1}', data_bar_rule)

# Create chart sheet
chart_ws = wb.create_sheet(title="Revenue & Profit Chart")

# Copy data to chart sheet for reference
for row_idx, row in df.iterrows():
    row_num = row_idx + 1
    chart_ws.cell(row=row_num, column=1, value=row['Month'])
    chart_ws.cell(row=row_num, column=2, value=row['Revenue'])
    chart_ws.cell(row=row_num, column=3, value=row['Profit'])

# Create clustered column chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Monthly Revenue and Profit"
chart.y_axis.title = "Amount ($)"
chart.x_axis.title = "Month"

# Data for chart
categories = Reference(chart_ws, min_col=1, min_row=2, max_row=len(df)+1)
revenue_data = Reference(chart_ws, min_col=2, min_row=1, max_row=len(df)+1)
profit_data = Reference(chart_ws, min_col=3, min_row=1, max_row=len(df)+1)

# Add series
revenue_series = Series(revenue_data, title="Revenue")
profit_series = Series(profit_data, title="Profit")

chart.append(revenue_series)
chart.append(profit_series)
chart.set_categories(categories)

# Add data labels
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

# Position chart
chart_ws.add_chart(chart, "E2")

# Save workbook
wb.save('kpi_report_formatted.xlsx')

print("\nKPI report generated successfully: kpi_report_formatted.xlsx")
print(f"Total rows: {len(df)}, Total columns: {len(df.columns)}")